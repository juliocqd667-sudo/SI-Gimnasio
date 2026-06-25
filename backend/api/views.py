from rest_framework import viewsets, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from datetime import timedelta
from core.models import CustomUser, Administrativo, Cliente, Instructor, Nutricionista
from finanzas.models import Promocion, Suscripcion, Pago, Comprobante
from actividades.models import Rutina, Ejercicio, DetalleRutina, Sala, Disciplina, Horario, Reserva, Seguimiento, Antecedentes
from . import serializers
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io

class CurrentUserView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        serializer = serializers.CustomUserSerializer(user)
        return Response(serializer.data)

# --- CORE ---
class CustomUserViewSet(viewsets.ModelViewSet):
    queryset = CustomUser.objects.all()
    serializer_class = serializers.CustomUserSerializer

    def perform_create(self, serializer):
        user = serializer.save()
        self._manage_profiles(user)

    def perform_update(self, serializer):
        user = serializer.save()
        self._manage_profiles(user)

    def _manage_profiles(self, user):
        if user.is_cliente:
            Cliente.objects.get_or_create(user=user)
        else:
            Cliente.objects.filter(user=user).delete()
             
        if user.is_instructor:
            Instructor.objects.get_or_create(user=user)
        else:
            Instructor.objects.filter(user=user).delete()
             
        if user.is_nutricionista:
            Nutricionista.objects.get_or_create(user=user)
        else:
            Nutricionista.objects.filter(user=user).delete()
             
        if user.is_administrativo:
            Administrativo.objects.get_or_create(user=user)
        else:
            Administrativo.objects.filter(user=user).delete()

class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = serializers.ClienteSerializer

class InstructorViewSet(viewsets.ModelViewSet):
    queryset = Instructor.objects.all()
    serializer_class = serializers.InstructorSerializer

class NutricionistaViewSet(viewsets.ModelViewSet):
    queryset = Nutricionista.objects.all()
    serializer_class = serializers.NutricionistaSerializer

class AdministrativoViewSet(viewsets.ModelViewSet):
    queryset = Administrativo.objects.all()
    serializer_class = serializers.AdministrativoSerializer

# --- FINANZAS ---
class PromocionViewSet(viewsets.ModelViewSet):
    queryset = Promocion.objects.all()
    serializer_class = serializers.PromocionSerializer

class SuscripcionViewSet(viewsets.ModelViewSet):
    queryset = Suscripcion.objects.all()
    serializer_class = serializers.SuscripcionSerializer

class PagoViewSet(viewsets.ModelViewSet):
    queryset = Pago.objects.all()
    serializer_class = serializers.PagoSerializer

    def perform_create(self, serializer):
        pago = serializer.save()
        cliente = pago.cliente
        suscripcion = pago.suscripcion
         
        days = 30
        plan = suscripcion.plan.lower()
        if 'anual' in plan or 'año' in plan: days = 365
        elif 'trimestral' in plan: days = 90
        elif 'semestral' in plan: days = 180
        elif 'semanal' in plan: days = 7
         
        fecha_ini = pago.fecha
        fecha_fin = fecha_ini + timedelta(days=days)
         
        cliente.suscripcion_activa = suscripcion.membresia
        cliente.fecha_ini_mem = fecha_ini
        cliente.fecha_fin_mem = fecha_fin
        cliente.save()
         
        Comprobante.objects.create(
            monto=pago.monto_total,
            fecha_ini_mem=fecha_ini,
            fecha_fin_mem=fecha_fin,
            pago=pago,
            cliente=cliente
        )

class ComprobanteViewSet(viewsets.ModelViewSet):
    queryset = Comprobante.objects.all()
    serializer_class = serializers.ComprobanteSerializer

# --- ACTIVIDADES ---
class RutinaViewSet(viewsets.ModelViewSet):
    queryset = Rutina.objects.all()
    serializer_class = serializers.RutinaSerializer

class EjercicioViewSet(viewsets.ModelViewSet):
    queryset = Ejercicio.objects.all()
    serializer_class = serializers.EjercicioSerializer

class DetalleRutinaViewSet(viewsets.ModelViewSet):
    queryset = DetalleRutina.objects.all()
    serializer_class = serializers.DetalleRutinaSerializer

class SalaViewSet(viewsets.ModelViewSet):
    queryset = Sala.objects.all()
    serializer_class = serializers.SalaSerializer

class DisciplinaViewSet(viewsets.ModelViewSet):
    queryset = Disciplina.objects.all()
    serializer_class = serializers.DisciplinaSerializer

class HorarioViewSet(viewsets.ModelViewSet):
    queryset = Horario.objects.all()
    serializer_class = serializers.HorarioSerializer

class ReservaViewSet(viewsets.ModelViewSet):
    queryset = Reserva.objects.all()
    serializer_class = serializers.ReservaSerializer

class SeguimientoViewSet(viewsets.ModelViewSet):
    queryset = Seguimiento.objects.all()
    serializer_class = serializers.SeguimientoSerializer

class AntecedentesViewSet(viewsets.ModelViewSet):
    queryset = Antecedentes.objects.all()
    serializer_class = serializers.AntecedentesSerializer

# Reportes
class ReporteUsuariosExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        # Headers
        headers = ['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Is Staff', 'Is Cliente', 'Is Instructor', 'Is Nutricionista', 'Is Administrativo']
        ws.append(headers)

        # Data
        usuarios = CustomUser.objects.all()
        for user in usuarios:
            ws.append([
                user.id,
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.is_staff,
                user.is_cliente,
                user.is_instructor,
                user.is_nutricionista,
                user.is_administrativo
            ])

        # Adjust column width
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Create HTTP response with Excel content
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
        wb.save(response)
        return response

class ReporteUsuariosPDFView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        elements.append(Paragraph("Reporte de Usuarios", styles['Title']))
        elements.append(Spacer(1, 12))

        # Table 
        data = [['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Is Staff', 'Is Cliente', 'Is Instructor', 'Is Nutricionista', 'Is Administrativo']]
        usuarios = CustomUser.objects.all()
        for user in usuarios:
            data.append([
                str(user.id),
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                str(user.is_staff),
                str(user.is_cliente),
                str(user.is_instructor),
                str(user.is_nutricionista),
                str(user.is_administrativo)
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)

        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=usuarios.pdf'
        return response

# --- REPORTES ---
class ReporteBaseExcel(APIView):
    permission_classes = [permissions.IsAuthenticated]
    model = None
    filename = 'reporte.xlsx'
    headers = []
    fields = []

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws.append(self.headers)
        for obj in self.model.objects.all():
            ws.append([getattr(obj, f) for f in self.fields])
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={self.filename}'
        wb.save(response)
        return response


class ReporteBasePDF(APIView):
    permission_classes = [permissions.IsAuthenticated]
    model = None
    filename = 'reporte.pdf'
    title = 'Reporte'
    headers = []
    fields = []

    def get(self, request):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(self.title, styles['Title']))
        elements.append(Spacer(1, 12))
        data = [self.headers]
        for obj in self.model.objects.all():
            data.append([str(getattr(obj, f)) for f in self.fields])
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename={self.filename}'
        return response


class ReportePagosExcelView(ReporteBaseExcel):
    model = Pago
    filename = 'pagos.xlsx'
    headers = ['ID', 'Cliente', 'Administrador', 'Membresía', 'Plan', 'Monto', 'Fecha', 'Método', 'Estado']
    fields = ['id', 'cliente', 'administrador', 'suscripcion', 'suscripcion', 'monto_total', 'fecha', 'metodo_pago', 'estado_pago']


class ReportePagosPDFView(ReporteBasePDF):
    model = Pago
    filename = 'pagos.pdf'
    title = 'Reporte de Pagos'
    headers = ['ID', 'Cliente', 'Administrador', 'Membresía', 'Plan', 'Monto', 'Fecha', 'Método', 'Estado']
    fields = ['id', 'cliente', 'administrador', 'suscripcion', 'suscripcion', 'monto_total', 'fecha', 'metodo_pago', 'estado_pago']


class ReporteSuscripcionesExcelView(ReporteBaseExcel):
    model = Suscripcion
    filename = 'suscripciones.xlsx'
    headers = ['ID', 'Membresía', 'Plan', 'Descripción', 'Precio']
    fields = ['id', 'membresia', 'plan', 'descripcion', 'precio']


class ReporteSuscripcionesPDFView(ReporteBasePDF):
    model = Suscripcion
    filename = 'suscripciones.pdf'
    title = 'Reporte de Suscripciones'
    headers = ['ID', 'Membresía', 'Plan', 'Descripción', 'Precio']
    fields = ['id', 'membresia', 'plan', 'descripcion', 'precio']


class ReportePromocionesExcelView(ReporteBaseExcel):
    model = Promocion
    filename = 'promociones.xlsx'
    headers = ['ID', 'Nombre', 'Tipo', 'Valor', 'Estado', 'Fecha Inicio', 'Fecha Fin', 'Descripción']
    fields = ['id', 'nombre', 'tipo', 'valor', 'estado', 'fecha_ini', 'fecha_fin', 'descripcion']


class ReportePromocionesPDFView(ReporteBasePDF):
    model = Promocion
    filename = 'promociones.pdf'
    title = 'Reporte de Promociones'
    headers = ['ID', 'Nombre', 'Tipo', 'Valor', 'Estado', 'Fecha Inicio', 'Fecha Fin', 'Descripción']
    fields = ['id', 'nombre', 'tipo', 'valor', 'estado', 'fecha_ini', 'fecha_fin', 'descripcion']


class ReporteRutinasExcelView(ReporteBaseExcel):
    model = Rutina
    filename = 'rutinas.xlsx'
    headers = ['ID', 'Nombre', 'Descripción']
    fields = ['id', 'nombre', 'descripcion']


class ReporteRutinasPDFView(ReporteBasePDF):
    model = Rutina
    filename = 'rutinas.pdf'
    title = 'Reporte de Rutinas'
    headers = ['ID', 'Nombre', 'Descripción']
    fields = ['id', 'nombre', 'descripcion']


class ReporteHorariosExcelView(ReporteBaseExcel):
    model = Horario
    filename = 'horarios.xlsx'
    headers = ['ID', 'Disciplina', 'Día', 'Hora Inicio', 'Hora Fin']
    fields = ['id', 'disciplina', 'dia', 'hora_inicial', 'hora_final']


class ReporteHorariosPDFView(ReporteBasePDF):
    model = Horario
    filename = 'horarios.pdf'
    title = 'Reporte de Horarios'
    headers = ['ID', 'Disciplina', 'Día', 'Hora Inicio', 'Hora Fin']
    fields = ['id', 'disciplina', 'dia', 'hora_inicial', 'hora_final']


class ReporteAntecedentesExcelView(ReporteBaseExcel):
    model = Antecedentes
    filename = 'antecedentes.xlsx'
    headers = ['ID', 'Cliente', 'Nutricionista', 'Fecha', 'Diagnóstico', 'Objetivo', 'Peso', 'Altura', 'IMC']
    fields = ['id', 'cliente', 'nutricionista', 'fecha', 'diagnostico', 'objetivo', 'peso', 'altura', 'imc']


class ReporteAntecedentesPDFView(ReporteBasePDF):
    model = Antecedentes
    filename = 'antecedentes.pdf'
    title = 'Reporte de Antecedentes'
    headers = ['ID', 'Cliente', 'Nutricionista', 'Fecha', 'Diagnóstico', 'Objetivo', 'Peso', 'Altura', 'IMC']
    fields = ['id', 'cliente', 'nutricionista', 'fecha', 'diagnostico', 'objetivo', 'peso', 'altura', 'imc']


class ReporteSeguimientosExcelView(ReporteBaseExcel):
    model = Seguimiento
    filename = 'seguimientos.xlsx'
    headers = ['ID', 'Cliente', 'Instructor', 'Rutina', 'Fecha', 'Próxima Consulta', 'Objetivo', 'Observaciones']
    fields = ['id', 'cliente', 'instructor', 'rutina', 'fecha', 'fecha_prox_consulta', 'objetivo', 'observaciones']


class ReporteSeguimientosPDFView(ReporteBasePDF):
    model = Seguimiento
    filename = 'seguimientos.pdf'
    title = 'Reporte de Seguimientos'
    headers = ['ID', 'Cliente', 'Instructor', 'Rutina', 'Fecha', 'Próxima Consulta', 'Objetivo', 'Observaciones']
    fields = ['id', 'cliente', 'instructor', 'rutina', 'fecha', 'fecha_prox_consulta', 'objetivo', 'observaciones']
